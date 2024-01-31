from jarvis.action.base_action import BaseAction
import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
class insert_line_chart_with_data(BaseAction):
    def __init__(self):
        self._description = "Insert a line chart into an Excel sheet using the provided data lists."

    def __call__(self,file_path, sheet_name, chart_title, x_title, y_title, x_data, y_data):
        """
        Insert a line chart into an Excel sheet using the provided data lists.

        Args:
        file_path (str): The path to the Excel file.
        sheet_name (str): The name of the sheet where the chart will be inserted.
        chart_title (str): The title of the chart.
        x_title (str): The title of the X-axis.
        y_title (str): The title of the Y-axis.
        x_data (list): A list of values for the X-axis (categories).
        y_data (list): A list of values for the Y-axis (values).

        Returns:
        None: The function modifies the file in place and does not return anything.
        """
        # Load the workbook and the sheet
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]

        # Write the x_data and y_data to the sheet
        max_row = len(x_data) + 1  # Adding 1 for the header row
        for index, value in enumerate(x_data, start=2):  # Starting from row 2 to skip header
            cell = sheet.cell(row=index, column=1)
            cell.value = value

        for index, value in enumerate(y_data, start=2):  # Assuming y_data goes in column 2
            cell = sheet.cell(row=index, column=2)
            cell.value = value

        # Create the chart object
        chart = LineChart()
        chart.title = chart_title
        chart.x_axis.title = x_title
        chart.y_axis.title = y_title

        # Add data to the chart
        data = Reference(sheet, min_col=2, min_row=1, max_row=max_row)  # Assumes data starts from row 2
        chart.add_data(data, titles_from_data=True)

        # Set categories (X-axis labels)
        cats = Reference(sheet, min_col=1, min_row=2, max_row=max_row)  # Assumes categories start from row 2
        chart.set_categories(cats)

        # Add the chart to the sheet
        sheet.add_chart(chart, "E15")  # Position the chart starting at cell E15

        # Save the workbook
        workbook.save(file_path)

# Example usage:
# res = insert_line_chart_with_data()(
#     file_path='path_to_your_excel_file.xlsx',
#     sheet_name='Sheet1',
#     chart_title='Daily chart',
#     x_title='Day',
#     y_title='Count',
#     x_data=['day 1', 'day 2', 'day 3', 'day 4', 'day 5', 'day 6', 'day 7', 'day 8', 'day 9', 'day 10'],
#     y_data=[112, 112, 151, 156, 10885, 147, 191, 144, 169, 168]
# )
# print(res)
