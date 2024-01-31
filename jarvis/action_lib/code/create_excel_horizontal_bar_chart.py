from jarvis.action.base_action import BaseAction
import os
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

class create_excel_horizontal_bar_chart(BaseAction):
    def __init__(self):
        self._description = "Create a horizontal bar chart in an Excel file."

    def __call__(self, excel_path, chart_title, x_col_name, y_col_name, sheet_name):
        """
        Create a horizontal bar chart in an Excel file.

        Parameters:
        excel_path (str): Path to the Excel file.
        chart_title (str): Title of the bar chart.
        x_col_name (str): Column name of the x-axis data.
        y_col_name (str): Column name of the y-axis data. 
        sheet_name (str): Name of the sheet in the workbook.
        """
        # Load the workbook and select the active worksheet
        wb = load_workbook(excel_path)
        sheet = wb[sheet_name]

        # Find the column numbers based on column names
        for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=1, values_only=True)):
            for j, cell_value in enumerate(row):
                if cell_value == x_col_name:
                    x_col = j + 1  # Convert to 1-based index
                if cell_value == y_col_name:
                    y_col = j + 1  # Convert to 1-based index

        # Determine the range of data for the chart
        min_row = 2
        max_row = min_row + sheet.max_row - 1

        # Create a new bar chart with horizontal bars, without axis titles and legend
        chart = BarChart()
        chart.type = 'bar'
        chart.style = 10
        chart.title = chart_title

        # Adjusting data references using column letters
        values = Reference(sheet, min_col=x_col, min_row=min_row, max_row=max_row)
        categories = Reference(sheet, min_col=y_col, min_row=min_row, max_row=max_row)

        # Adding data to the chart
        chart.add_data(values, titles_from_data=False)
        chart.set_categories(categories)

        # Removing the legend
        chart.legend = None

        # Place the chart on the sheet
        sheet.add_chart(chart, "E2")

        # Save the workbook with the new chart
        wb.save(excel_path)

        print("horizontal bar chart is successfully created!")

# Example usage:
# res = create_excel_horizontal_bar_chart()("/home/heroding/桌面/Jarvis/tasks/SheetTasks/sheets/StockChange.xlsx", "Value of Stock on Dec 31", "Value of Stock on Dec 31", "Stock", "Sheet1")
# print(res)
