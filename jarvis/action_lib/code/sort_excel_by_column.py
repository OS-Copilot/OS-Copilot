from jarvis.action.base_action import BaseAction
import openpyxl
from typing import Tuple
class sort_excel_by_column(BaseAction):
    def __init__(self):
        self._description = "Sorts an Excel sheet based on a specified column in ascending or descending order."

    def __call__(self,file_path: str, sheet_name: str, sort_by: str, order: str = 'asc') -> None:
        """
        Sorts an Excel sheet based on a specified column in ascending or descending order.

        Args:
        file_path (str): Path to the Excel file.
        sheet_name (str): Name of the sheet to be sorted.
        sort_by (str): The column header used for sorting.
        order (str): Sorting order, 'asc' for ascending or 'dec' for descending. Default is 'asc'.

        Returns:
        None: The function modifies the file in place and does not return anything.
        """
        # Load the workbook and select the specified sheet
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]

        # Finding the index of the sort column
        header_row = next(sheet.iter_rows(values_only=True))
        sort_column_idx = header_row.index(sort_by) + 1

        # Convert the worksheet to a list of tuples
        data = list(sheet.iter_rows(values_only=True))

        # Sort the data
        sorted_data = sorted(data[1:], key=lambda x: x[sort_column_idx-1] if x[sort_column_idx-1] is not None else "")

        # Determine the sort order
        if order.lower() == 'dec':
            sorted_data.reverse()

        # Write back the sorted data to the worksheet
        for row_idx, row in enumerate(sorted_data, start=2):
            for col_idx, value in enumerate(row, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

        # Save the workbook
        workbook.save(file_path)

# Example usage
# print(sort_excel()('/path/to/your/file.xlsx', 'YourSheetName', 'YourSortColumn', 'asc'))