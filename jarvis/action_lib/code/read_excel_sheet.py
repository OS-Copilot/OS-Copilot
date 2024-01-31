from jarvis.action.base_action import BaseAction
import os
from collections import Counter
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from typing import List, Any, Optional
class read_excel_sheet(BaseAction):
    def __init__(self):
        self._description = "Reads all data from a specified sheet in an Excel file."

    def __call__(self,file_path: str, sheet_name: Optional[str] = None) -> List[List[Any]]:
        """
        Reads all data from a specified sheet in an Excel file.

        Args:
        file_path (str): Path to the Excel file.
        sheet_name (Optional[str]): Name of the sheet to be read. If None, reads the first sheet.

        Returns:
        List[List[Any]]: A 2D list containing all data from the specified Excel sheet.

        """
        try:
            # Load the workbook and the specified sheet
            workbook = openpyxl.load_workbook(file_path, data_only=True)

            # If sheet_name is not provided, use the first sheet
            if sheet_name is None:
                sheet = workbook.active
            else:
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                else:
                    raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")

            # Reading all data from the sheet
            data = [[cell.value for cell in row] for row in sheet.iter_rows()]

            return data

        except FileNotFoundError:
            raise FileNotFoundError(f"The file '{file_path}' was not found.")
        except InvalidFileException:
            raise InvalidFileException(f"The file '{file_path}' is not a valid Excel file.")

# Example usage:
# res = extract_excel_content()("/home/heroding/桌面/Jarvis/tasks/SheetTasks/sheets/BoomerangSales.xlsx","Sheet1")
# print(res)
