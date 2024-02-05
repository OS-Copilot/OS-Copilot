from jarvis.action.base_action import BaseAction
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List, Optional
class set_column_cells_format(BaseAction):
    def __init__(self):
        self._description = "Modify the style of specified cells in a column of an Excel sheet."

    def __call__(self,file_path: str, sheet_name: str, column_name: str, rows: List[int],
                              font_name: Optional[str] = None, font_size: Optional[float] = None,
                              font_color: Optional[str] = None, fill_color: Optional[str] = None,
                              bold: Optional[bool] = None, italic: Optional[bool] = None,
                              underline: Optional[bool] = None, horizontal_alignment: Optional[str] = None) -> None:
        """
        Modify the style of specified cells in a column of an Excel sheet.

        Parameters:
        file_path (str): The path to the Excel file.
        sheet_name (str): The name of the sheet where the cells are located.
        column_name (str): The name of the column to modify (e.g., 'Gross Profit').
        rows (List[int]): A list of row numbers to apply the format to.
        font_name (Optional[str]): The name of the font to apply.
        font_size (Optional[float]): The size of the font to apply.
        font_color (Optional[str]): The color of the font in hex format (e.g., 'FF0000' for red).
        fill_color (Optional[str]): The fill color of the cell in hex format.
        bold (Optional[bool]): Set to True to make text bold.
        italic (Optional[bool]): Set to True to make text italic.
        underline (Optional[bool]): Set to True to apply single underline.
        horizontal_alignment (Optional[str]): The horizontal alignment of the text (e.g., 'center').

        Returns:
        None: The function modifies the file in place and does not return anything.
        """

        # Load the workbook and the sheet
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]

        # Find the column number for the given column name
        column_number = None
        for cell in sheet[1]:
            if cell.value == column_name:
                column_number = cell.column
                break
        if column_number is None:
            raise ValueError(f"Column name '{column_name}' not found.")

        # Apply the formatting to the specified rows in the found column
        for row in rows:
            cell = sheet.cell(row=row, column=column_number)
            cell.font = Font(name=font_name, size=font_size, color=font_color, bold=bold, italic=italic, underline='single' if underline else None)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid") if fill_color else None
            cell.alignment = Alignment(horizontal=horizontal_alignment) if horizontal_alignment else None

        # Save the workbook
        workbook.save(file_path)

# Example usage:
# res = set_column_cells_format()(
#     file_path='path_to_your_excel_file.xlsx',
#     sheet_name='Sheet1',
#     column_name='Target Column',
#     rows=[6, 7, 8, 9],
#     font_name='Calibri',
#     font_size=12,
#     font_color='FF0000',
#     fill_color='00FF00',
#     bold=True,
#     italic=False,
#     underline=True,
#     horizontal_alignment='center'
# )
# print(res)
